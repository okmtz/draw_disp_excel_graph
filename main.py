import glob
import re
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series

def main():
    # ディレクトリ内の.excelを読み込み
    files = glob.glob("./inputs/*.xlsx")
    for file in files:
        print("#############################################")
        print(f'loading file {file}')
        print("#############################################")
        # Excelファイルの読み込み
        work_book = load_workbook(file)

        print("##############################################")
        print('now executing')
        print("##############################################")
        sheet_name_re = re.search(r'A\d+-T\d+', file)
        sheet_name = sheet_name_re.group()
        sheet = work_book[sheet_name]
        # 散布図を定義
        chart = ScatterChart()
        # y, xデータの範囲を選択
        max_row_num = sheet.max_row
        y = Reference(work_book[sheet_name], min_col=10, max_col=10, min_row=13, max_row=max_row_num)
        x = Reference(work_book[sheet_name], min_col=9, max_col=9, min_row=13, max_row=max_row_num)
        series = Series(y, x)
        # デフォルトだと線で表示されるので、線を消す
        series.graphicalProperties.line.noFill = True
        # マーカーを表示させる
        series.marker.symbol = "auto"
        # 散布図として定義したchartへデータを渡す
        chart.series.append(series)
        # グラフ描画用のシートを追加
        work_book.create_sheet(title='disp_graph')
        # 指定したセルにグラフを表示する
        work_book['disp_graph'].add_chart(chart, "A6")
        work_book.save(f'./outputs/{sheet_name}-with-disp.xlsx')
        print("###############################################")
        print(f'file {sheet_name} is done.')
        print("################################################")
        


if __name__ == '__main__':
    main()
